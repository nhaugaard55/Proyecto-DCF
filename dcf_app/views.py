import io
import re
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, cast
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.core.cache import cache
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from xhtml2pdf import pisa

from .models import AnalysisRecord, WatchlistItem

from dcf_core.DCF_Main import ejecutar_dcf
from dcf_core.business_cycle import get_business_cycle_phase
from dcf_core.company_stage import detect_company_stage, STAGE_META
from dcf_core.multi_model_valuation import run_all_models
from dcf_core.search import CompanySearchResult, search_companies


_SYMBOL_PATTERN = re.compile(r"^\s*([A-Za-z0-9.\-:]+)")

_DCF_CACHE_TTL = 600  # 10 minutos
_AUTO_FUENTE = "auto"


def _cached_ejecutar_dcf(ticker: str) -> dict:
    """Ejecuta DCF automático con caché de 10 minutos por ticker."""
    cache_key = f"dcf_result_auto_{ticker}"
    cached = cache.get(cache_key)
    if cached is not None:
        return cached
    resultado = ejecutar_dcf(ticker, "auto", _AUTO_FUENTE)
    cache.set(cache_key, resultado, _DCF_CACHE_TTL)
    return resultado


def _resolver_ticker(raw_ticker: str, raw_query: str) -> str:
    """Normaliza el símbolo a partir de la selección o del texto libre."""

    ticker_limpio = (raw_ticker or "").strip().upper()
    if ticker_limpio:
        return ticker_limpio

    query_limpio = (raw_query or "").strip()
    if not query_limpio:
        return ""

    match = _SYMBOL_PATTERN.match(query_limpio)
    potencial = match.group(1).upper() if match else ""

    if potencial and " " not in potencial:
        return potencial

    coincidencias = search_companies(query_limpio, limit=1)
    if coincidencias:
        return coincidencias[0].symbol.upper()

    return potencial or query_limpio.upper()


def _to_decimal(value, *, places: int | None = 4):
    if value in (None, "", "N/D"):
        return None
    if isinstance(value, Decimal):
        dec_value = value
    else:
        try:
            dec_value = Decimal(str(value))
        except (InvalidOperation, TypeError, ValueError):
            return None
    if places is None:
        return dec_value
    quant = Decimal("1." + ("0" * places))
    return dec_value.quantize(quant, rounding=ROUND_HALF_UP)


def _clean_numeric(value):
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value))
        except (TypeError, ValueError):
            return None


def _extract_chart_series(entries):
    labels = []
    values = []
    if not entries:
        return labels, values

    for item in entries:
        if isinstance(item, dict):
            labels.append(str(item.get('anio', '')))
            values.append(_clean_numeric(item.get('valor')))
        else:
            labels.append(str(getattr(item, 'anio', '')))
            values.append(_clean_numeric(getattr(item, 'valor', None)))
    return labels, values


def _build_chart_data(resultado):
    if not resultado:
        return {
            'has_resultado': False,
            'fcf_historico_labels': [],
            'fcf_historico_series': [],
            'fcf_proyectado_labels': [],
            'fcf_proyectado_series': [],
        }

    if isinstance(resultado, dict):
        historico_entries = resultado.get('fcf_historico')
        proyectado_entries = resultado.get('fcf_proyectado')
    else:
        historico_entries = getattr(resultado, 'fcf_historico', None)
        proyectado_entries = getattr(resultado, 'fcf_proyectado', None)

    historico_labels, historico_series = _extract_chart_series(historico_entries)
    proyectado_labels, proyectado_series = _extract_chart_series(proyectado_entries)

    has_data = bool(historico_labels or proyectado_labels)
    return {
        'has_resultado': has_data,
        'fcf_historico_labels': historico_labels,
        'fcf_historico_series': historico_series,
        'fcf_proyectado_labels': proyectado_labels,
        'fcf_proyectado_series': proyectado_series,
    }


NEWS_PAGE_SIZE = 6
RECENT_HISTORY_VISIBLE_LIMIT = 5
RECENT_HISTORY_FETCH_LIMIT = 25


def _guardar_analisis(
    *,
    ticker: str,
    company_name: str,
    company_exchange: str,
    resultado: dict[str, Any] | None,
):
    if not ticker or not isinstance(resultado, dict):
        return None

    nombre_empresa = (company_name or resultado.get("nombre") or ticker).strip()
    sector = (resultado.get("sector") or "").strip()
    fuente_utilizada = (resultado.get("fuente_datos") or "").strip()
    metodo = (
        (resultado.get("datos_empresa") or {}).get("metodo_crecimiento_codigo")
        or AnalysisRecord.METODO_CAGR
    )

    valor_intrinseco = _to_decimal(resultado.get("valor_intrinseco"), places=4)
    precio_actual = _to_decimal(resultado.get("precio_actual"), places=4)
    diferencia_pct = _to_decimal(resultado.get("diferencia_pct"), places=2)

    ventana_reciente = timezone.now() - timedelta(minutes=5)
    duplicado = AnalysisRecord.objects.filter(
        ticker=ticker,
        metodo=metodo,
        fuente_utilizada=fuente_utilizada,
        valor_intrinseco=valor_intrinseco,
        precio_actual=precio_actual,
        diferencia_pct=diferencia_pct,
        created_at__gte=ventana_reciente,
    ).exists()

    if duplicado:
        return None

    return AnalysisRecord.objects.create(
        ticker=ticker,
        company_name=nombre_empresa,
        company_exchange=company_exchange,
        sector=sector,
        metodo=metodo,
        fuente_solicitada=_AUTO_FUENTE,
        fuente_utilizada=fuente_utilizada,
        valor_intrinseco=valor_intrinseco,
        precio_actual=precio_actual,
        diferencia_pct=diferencia_pct,
        estado=(resultado.get("estado") or "").strip(),
    )


def _serialize_news_item(item: dict) -> dict:
    fecha = item.get("fecha")
    fecha_iso = None
    fecha_display = None
    if isinstance(fecha, datetime):
        fecha_local = timezone.localtime(fecha) if timezone.is_aware(fecha) else fecha
        fecha_iso = fecha_local.isoformat()
        fecha_display = fecha_local.strftime("%d/%m/%Y %H:%M")

    return {
        "titulo": item.get("titulo"),
        "fuente": item.get("fuente"),
        "resumen": item.get("resumen"),
        "url": item.get("url"),
        "imagen": item.get("imagen"),
        "fecha_iso": fecha_iso,
        "fecha_display": fecha_display,
    }


def _parse_page(value: str | None, default: int = 1) -> int:
    if value is None:
        return default
    value = value.strip()
    if not value:
        return default
    try:
        numero = int(value)
        return numero if numero > 0 else default
    except (TypeError, ValueError):
        return default


def landing(request):
    return render(request, 'landing.html')


_TICKER_STRIP_SYMBOLS = ["AAPL", "MSFT", "NVDA", "GOOGL", "AMZN", "META", "TSLA", "BRK-B", "JPM", "V"]
_TICKER_STRIP_CACHE_TTL = 300  # 5 minutos


@require_GET
def ticker_strip_view(request):
    """Devuelve precios y variación diaria de los tickers del strip (JSON, caché 5 min)."""
    import yfinance as yf

    cache_key = "ticker_strip_data"
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse({"tickers": cached})

    try:
        data = yf.download(
            _TICKER_STRIP_SYMBOLS,
            period="2d",
            interval="1d",
            progress=False,
            auto_adjust=True,
        )
        results = []
        close = data["Close"] if "Close" in data.columns else data.xs("Close", axis=1, level=0)
        for sym in _TICKER_STRIP_SYMBOLS:
            display = sym.replace("-", ".")
            try:
                series = close[sym].dropna()
                if len(series) < 2:
                    continue
                prev, last = float(series.iloc[-2]), float(series.iloc[-1])
                chg = (last - prev) / prev * 100
                results.append({
                    "sym": display,
                    "price": f"{last:.2f}",
                    "chg": f"{chg:+.2f}%",
                    "up": chg >= 0,
                })
            except Exception:
                continue

        cache.set(cache_key, results, _TICKER_STRIP_CACHE_TTL)
        return JsonResponse({"tickers": results})
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


def dcf_view(request):
    resultado = None
    error = None
    ticker = ""
    valor_busqueda = request.GET.get("company_query", "").strip()
    company_name = request.GET.get("company_name", "").strip()
    company_exchange = request.GET.get("company_exchange", "").strip()

    page_number = _parse_page(request.GET.get("page"))

    if request.method == "POST":
        valor_busqueda = request.POST.get("company_query", "").strip()
        company_name = request.POST.get("company_name", "").strip()
        company_exchange = request.POST.get("company_exchange", "").strip()
        ticker_resuelto = _resolver_ticker(request.POST.get("ticker", ""), valor_busqueda)

        if ticker_resuelto:
            query_params = {
                "ticker": ticker_resuelto,
                "company_query": valor_busqueda,
                "company_name": company_name,
                "company_exchange": company_exchange,
                "page": 1,
            }
            return redirect(f"{reverse('home')}?{urlencode(query_params)}")

        error = "Por favor ingresá un ticker válido."

    ticker = request.GET.get("ticker", "").strip().upper()

    if ticker:
        try:
            resultado = _cached_ejecutar_dcf(ticker)
        except Exception as exc:
            error = f"Ocurrió un error al analizar el ticker: {exc}"
            resultado = None
        else:
            _guardar_analisis(
                ticker=ticker,
                company_name=company_name,
                company_exchange=company_exchange,
                resultado=resultado,
            )

    company_stage = None
    multi_model = None
    if resultado and isinstance(resultado, dict):
        try:
            company_stage = detect_company_stage(ticker, resultado)
        except Exception:
            company_stage = None

        try:
            stage_num = (company_stage or {}).get("stage", 4)
            wacc_val = (resultado.get("metricas") or {}).get("wacc") or 0.08
            multi_model = run_all_models(ticker, resultado, stage_num, wacc_val)
        except Exception:
            multi_model = None

    chart_data = _build_chart_data(resultado)

    if isinstance(resultado, dict):
        raw_news = resultado.get("noticias") or []
    else:
        raw_news = getattr(resultado, "noticias", []) or []
    try:
        news_list: list[dict[str, Any]] = list(raw_news)
    except TypeError:
        news_list = []
    news_total = len(news_list)
    total_pages = max(1, (news_total + NEWS_PAGE_SIZE - 1) // NEWS_PAGE_SIZE) if news_total else 1
    if page_number > total_pages:
        page_number = total_pages
    if page_number < 1:
        page_number = 1

    base_query = request.GET.copy()
    if "page" in base_query:
        del base_query["page"]
    base_query_string = base_query.urlencode()

    news_payload = [_serialize_news_item(item) for item in news_list]
    recent_records_queryset = AnalysisRecord.objects.all()
    recent_records = list(recent_records_queryset[:RECENT_HISTORY_FETCH_LIMIT])

    tradingview_symbol = ticker
    if company_exchange and ticker:
        tradingview_symbol = f"{company_exchange.upper()}:{ticker}"

    in_watchlist = WatchlistItem.objects.filter(ticker=ticker).exists() if ticker else False
    precio_historico = (resultado or {}).get("precio_historico") if resultado else None

    context = {
        "in_watchlist": in_watchlist,
        "precio_historico": precio_historico,
        "multi_model": multi_model,
        "resultado": resultado,
        "error": error,
        "ticker": ticker,
        "search_value": valor_busqueda or ticker,
        "company_name": company_name,
        "company_exchange": company_exchange,
        "tradingview_symbol": tradingview_symbol,
        "chart_data": chart_data,
        "news_data": news_payload,
        "news_total": news_total,
        "news_page_size": NEWS_PAGE_SIZE,
        "news_initial_page": page_number,
        "news_base_query": base_query_string,
        "news_config": {
            "page_size": NEWS_PAGE_SIZE,
            "initial_page": page_number,
            "base_query": base_query_string,
        },
        "recent_records": recent_records,
        "recent_records_limit": RECENT_HISTORY_VISIBLE_LIMIT,
        "company_stage": company_stage,
        "stage_labels": ["Startup", "Hyper Growth", "Break Even", "Op. Leverage", "Cap. Return", "Decline"],
        "all_stages": [
            {"stage": k, "nombre": v["nombre"], "descripcion_breve": v["descripcion_breve"], "color": v["color"]}
            for k, v in STAGE_META.items()
        ],
    }

    return render(request, "dcf_app/index.html", context)


def _render_pdf(template_name: str, context: dict) -> bytes | None:
    html = render_to_string(template_name, context)
    output = io.BytesIO()
    pdf_result = cast(Any, pisa.CreatePDF(html, dest=output, encoding="UTF-8"))
    if getattr(pdf_result, "err", 0):
        return None
    return output.getvalue()


def dcf_pdf_view(request):
    ticker = request.GET.get("ticker", "").strip().upper()

    if not ticker:
        return HttpResponse("Ticker inválido", status=400)

    try:
        resultado = _cached_ejecutar_dcf(ticker)
    except Exception as exc:
        return HttpResponse(f"No se pudo generar el informe: {exc}", status=500)

    if not resultado:
        return HttpResponse("No hay datos suficientes para generar el informe", status=404)

    context = {
        "resultado": resultado,
        "ticker": ticker,
        "generado": timezone.now(),
    }

    pdf_bytes = _render_pdf("dcf_app/pdf_report.html", context)
    if pdf_bytes is None:
        return HttpResponse("Error generando PDF", status=500)

    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="DCF_{ticker}.pdf"'
    return response


_BUSINESS_CYCLE_CACHE_TTL = 600  # 10 minutos


@require_GET
def business_cycle_view(request):
    """Devuelve la fase del ciclo económico detectada (JSON)."""
    cache_key = "business_cycle_phase"
    cached = cache.get(cache_key)
    if cached is not None:
        return JsonResponse(cached)

    try:
        data = get_business_cycle_phase()
        cache.set(cache_key, data, _BUSINESS_CYCLE_CACHE_TTL)
        return JsonResponse(data)
    except Exception as exc:
        return JsonResponse({"error": str(exc)}, status=500)


@require_GET
def search_companies_view(request):
    """Devuelve coincidencias de tickers basadas en el texto ingresado."""

    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse({"results": []})

    try:
        limit = int(request.GET.get("limit", 10))
    except (TypeError, ValueError):
        limit = 10

    limit = max(1, min(limit, 20))

    coincidencias = search_companies(query, limit=limit)

    def _serializar(item: CompanySearchResult) -> dict:
        return {
            "symbol": item.symbol,
            "name": item.name,
            "exchange": item.exchange,
            "type": item.asset_type,
        }

    return JsonResponse({"results": [_serializar(item) for item in coincidencias]})


# ---------------------------------------------------------------------------
# Watchlist
# ---------------------------------------------------------------------------

def watchlist_view(request):
    """Página principal de la watchlist."""
    items = WatchlistItem.objects.all()
    return render(request, "dcf_app/watchlist.html", {"watchlist": items})


@require_POST
def watchlist_toggle(request):
    """Agrega o quita un ticker de la watchlist (JSON)."""
    ticker = (request.POST.get("ticker") or "").strip().upper()
    company_name = (request.POST.get("company_name") or "").strip()
    company_exchange = (request.POST.get("company_exchange") or "").strip()

    if not ticker:
        return JsonResponse({"error": "Ticker requerido"}, status=400)

    item = WatchlistItem.objects.filter(ticker=ticker).first()
    if item:
        item.delete()
        return JsonResponse({"action": "removed", "ticker": ticker})
    else:
        WatchlistItem.objects.create(
            ticker=ticker,
            company_name=company_name,
            company_exchange=company_exchange,
        )
        return JsonResponse({"action": "added", "ticker": ticker})


@require_GET
def watchlist_status(request):
    """Devuelve si un ticker está en la watchlist."""
    ticker = (request.GET.get("ticker") or "").strip().upper()
    if not ticker:
        return JsonResponse({"in_watchlist": False})
    in_watchlist = WatchlistItem.objects.filter(ticker=ticker).exists()
    return JsonResponse({"in_watchlist": in_watchlist, "ticker": ticker})


# ---------------------------------------------------------------------------
# Comparar empresas
# ---------------------------------------------------------------------------

def comparar_view(request):
    """Muestra un análisis comparativo de dos tickers."""
    ticker_a = (request.GET.get("ticker_a") or "").strip().upper()
    ticker_b = (request.GET.get("ticker_b") or "").strip().upper()

    resultado_a = resultado_b = error_a = error_b = None

    if ticker_a:
        try:
            resultado_a = _cached_ejecutar_dcf(ticker_a)
        except Exception as exc:
            error_a = str(exc)

    if ticker_b:
        try:
            resultado_b = _cached_ejecutar_dcf(ticker_b)
        except Exception as exc:
            error_b = str(exc)

    context = {
        "ticker_a": ticker_a,
        "ticker_b": ticker_b,
        "resultado_a": resultado_a,
        "resultado_b": resultado_b,
        "error_a": error_a,
        "error_b": error_b,
    }
    return render(request, "dcf_app/comparar.html", context)


# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def dcf_excel_view(request):
    """Genera y descarga un .xlsx con el análisis DCF."""
    ticker = request.GET.get("ticker", "").strip().upper()

    if not ticker:
        return HttpResponse("Ticker inválido", status=400)

    try:
        resultado = _cached_ejecutar_dcf(ticker)
    except Exception as exc:
        return HttpResponse(f"Error al obtener datos: {exc}", status=500)

    wb = openpyxl.Workbook()

    # ---- Hoja 1: Resumen ----
    ws = wb.active
    ws.title = "Resumen"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = _hex_fill("2563EB")
    title_font = Font(bold=True, size=13)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20

    ws["A1"] = f"Análisis DCF — {resultado.get('nombre', ticker)}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:B1")

    rows = [
        ("Campo", "Valor"),
        ("Ticker", ticker),
        ("Empresa", resultado.get("nombre")),
        ("Sector", resultado.get("sector")),
        ("Precio actual", resultado.get("precio_actual")),
        ("Valor intrínseco", resultado.get("valor_intrinseco")),
        ("Diferencia %", resultado.get("diferencia_pct")),
        ("Estado", resultado.get("estado")),
        ("Fuente de datos", resultado.get("fuente_datos_descripcion")),
        ("Método", resultado.get("datos_empresa", {}).get("metodo_crecimiento")),
        ("WACC %", resultado.get("metricas", {}).get("wacc_pct")),
        ("Crecimiento %", resultado.get("metricas", {}).get("crecimiento_pct")),
        ("Beta", resultado.get("datos_empresa", {}).get("beta")),
    ]

    for i, (label, value) in enumerate(rows, start=2):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        if i == 2:
            ws[f"A{i}"].font = header_font
            ws[f"A{i}"].fill = header_fill
            ws[f"B{i}"].font = header_font
            ws[f"B{i}"].fill = header_fill

    # ---- Hoja 2: FCF Histórico y Proyectado ----
    ws2 = wb.create_sheet("FCF")
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 20
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 20

    ws2["A1"] = "FCF Histórico (B$)"
    ws2["A1"].font = Font(bold=True)
    ws2["D1"] = "FCF Proyectado (B$)"
    ws2["D1"].font = Font(bold=True)
    ws2["A2"], ws2["B2"] = "Año", "Valor"
    ws2["D2"], ws2["E2"] = "Año", "Valor"
    for cell in [ws2["A2"], ws2["B2"], ws2["D2"], ws2["E2"]]:
        cell.font = header_font
        cell.fill = header_fill

    for i, entry in enumerate(resultado.get("fcf_historico") or [], start=3):
        ws2[f"A{i}"] = entry.get("anio")
        ws2[f"B{i}"] = entry.get("valor")

    for i, entry in enumerate(resultado.get("fcf_proyectado") or [], start=3):
        ws2[f"D{i}"] = entry.get("anio")
        ws2[f"E{i}"] = entry.get("valor")

    # ---- Hoja 3: Escenarios ----
    escenarios = resultado.get("escenarios")
    if escenarios:
        ws3 = wb.create_sheet("Escenarios")
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 16
        ws3.column_dimensions["C"].width = 16
        ws3.column_dimensions["D"].width = 16
        headers = ["Campo", "Bear (Pesimista)", "Base", "Bull (Optimista)"]
        for col, h in enumerate(headers, start=1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        campos = [
            ("Crecimiento %", "tasa_crecimiento_pct"),
            ("Valor intrínseco", "valor_intrinseco"),
            ("Diferencia %", "diferencia_pct"),
            ("Estado", "estado"),
        ]
        for row_i, (label, key) in enumerate(campos, start=2):
            ws3.cell(row=row_i, column=1, value=label).font = Font(bold=True)
            for col_i, esc in enumerate(["bear", "base", "bull"], start=2):
                ws3.cell(row=row_i, column=col_i, value=(escenarios.get(esc) or {}).get(key))

    # ---- Hoja 4: Tabla de sensibilidad ----
    tabla = resultado.get("tabla_sensibilidad")
    if tabla:
        ws4 = wb.create_sheet("Sensibilidad")
        ws4["A1"] = "Valor intrínseco por acción (WACC × Crecimiento)"
        ws4["A1"].font = Font(bold=True, size=11)
        ws4.merge_cells(f"A1:{chr(65 + len(tabla['crecimientos']))}1")

        ws4["A2"] = "WACC \\ Crec."
        ws4["A2"].font = header_font
        ws4["A2"].fill = header_fill
        for col_i, g in enumerate(tabla["crecimientos"], start=2):
            cell = ws4.cell(row=2, column=col_i, value=f"{g}%")
            cell.font = header_font
            cell.fill = header_fill

        precio = tabla.get("precio_actual") or 0
        for row_i, (w, row_vals) in enumerate(zip(tabla["waccs"], tabla["matrix"]), start=3):
            ws4.cell(row=row_i, column=1, value=f"{w}%").font = Font(bold=True)
            for col_i, val in enumerate(row_vals, start=2):
                cell = ws4.cell(row=row_i, column=col_i, value=val)
                if val is not None and precio:
                    if val > precio * 1.1:
                        cell.fill = _hex_fill("DCFCE7")
                    elif val < precio * 0.9:
                        cell.fill = _hex_fill("FEE2E2")
                    else:
                        cell.fill = _hex_fill("FEF9C3")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="DCF_{ticker}.xlsx"'
    wb.save(response)
    return response
